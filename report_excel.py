"""
Excel report generator — ESL Safety Observations
Columns: Ref No | Date | Observation | Area | Responsibility | Status | Target Date | Picture
"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── BRAND COLORS ────────────────────────────────────────────────────────────
C_ORANGE  = "F97316"
C_BLUE    = "3B82F6"
C_GREEN   = "22C55E"
C_YELLOW  = "EAB308"
C_RED     = "EF4444"
C_DARK    = "0D1117"
C_SURFACE = "111827"
C_MUTED   = "9CA3AF"
C_WHITE   = "FFFFFF"
C_TEXT    = "E2E8F0"

STATUS_MAP = {
    "open":             ("EF4444", "FEE2E2"),
    "partially closed": ("EAB308", "FEF9C3"),
    "in progress":      ("EAB308", "FEF9C3"),
    "closed":           ("22C55E", "DCFCE7"),
}

HEADERS    = ["Ref No", "Date", "Observation", "Area",
              "Responsibility for Closure", "Status", "Target Date", "Picture"]
COL_WIDTHS = [14, 16, 55, 28, 30, 20, 14, 16]


def _border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(h): return PatternFill("solid", fgColor=h.lstrip("#"))


def generate_excel(observations: list[dict]) -> bytes:
    wb = Workbook()

    # ── SHEET 1: Observations ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Observations"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value     = "ESL STEEL LIMITED  —  Safety Observations Report"
    tc.fill      = _fill(C_DARK)
    tc.font      = Font(bold=True, size=14, color=C_TEXT, name="Segoe UI")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38

    ws.merge_cells("A2:H2")
    sc = ws["A2"]
    sc.value     = (f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
                    f"     |     Total Observations: {len(observations)}")
    sc.fill      = _fill(C_SURFACE)
    sc.font      = Font(size=10, color=C_MUTED, name="Segoe UI")
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Headers
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.fill      = _fill(C_ORANGE)
        c.font      = Font(bold=True, size=11, color=C_WHITE, name="Segoe UI")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border()
    ws.row_dimensions[3].height = 30

    # Data
    for r, obs in enumerate(observations, 4):
        status       = obs.get("status", "Open").strip().lower()
        s_txt, s_bg  = STATUS_MAP.get(status, ("374151", "F9FAFB"))
        row_bg       = "F0F4FF" if r % 2 == 0 else "FFFFFF"
        image_url    = obs.get("image_url", "").strip()

        vals = [
            obs.get("ref_no",       ""),
            obs.get("datetime",     ""),
            obs.get("observation",  ""),
            obs.get("area",         ""),
            obs.get("responsible",  ""),
            obs.get("status",       ""),
            obs.get("target_date",  ""),
            "📷 View Photo" if image_url else "No Photo",
        ]

        CENTER_COLS = {1, 2, 6, 7, 8}

        for c, val in enumerate(vals, 1):
            cell           = ws.cell(row=r, column=c, value=val)
            cell.border    = _border()
            cell.alignment = Alignment(
                vertical="center", wrap_text=True,
                horizontal="center" if c in CENTER_COLS else "left"
            )

            if c == 6:                      # Status
                cell.fill = _fill(s_bg)
                cell.font = Font(bold=True, size=11, color=s_txt, name="Segoe UI")
            elif c == 8 and image_url:      # Picture hyperlink
                cell.fill      = _fill(row_bg)
                cell.hyperlink = image_url
                cell.font      = Font(bold=True, size=11, color=C_BLUE,
                                      underline="single", name="Segoe UI")
            else:
                cell.fill = _fill(row_bg)
                cell.font = Font(size=11, name="Segoe UI")

        ws.row_dimensions[r].height = 22

    # Column widths + freeze + filter
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = f"A3:{get_column_letter(len(HEADERS))}3"
    ws.freeze_panes    = "A4"

    # ── SHEET 2: Summary ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.sheet_view.showGridLines = False

    total     = len(observations)
    open_c    = sum(1 for o in observations if o.get("status","").lower() == "open")
    partial_c = sum(1 for o in observations if "partial" in o.get("status","").lower()
                                            or "progress" in o.get("status","").lower())
    closed_c  = sum(1 for o in observations if o.get("status","").lower() == "closed")
    rate      = f"{closed_c / total * 100:.1f}%" if total else "N/A"

    ws2.merge_cells("B2:G2")
    t = ws2["B2"]
    t.value = "ESL STEEL LIMITED — Safety Summary"
    t.fill  = _fill(C_DARK)
    t.font  = Font(bold=True, size=15, color=C_TEXT, name="Segoe UI")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 38

    ws2.merge_cells("B3:G3")
    d = ws2["B3"]
    d.value = f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}"
    d.fill  = _fill(C_SURFACE)
    d.font  = Font(size=10, color=C_MUTED, name="Segoe UI")
    d.alignment = Alignment(horizontal="center")

    cards = [
        ("Total",            str(total),     C_BLUE,   "DBEAFE"),
        ("Open",             str(open_c),    C_RED,    "FEE2E2"),
        ("Partially Closed", str(partial_c), C_YELLOW, "FEF9C3"),
        ("Closed",           str(closed_c),  C_GREEN,  "DCFCE7"),
        ("Closure Rate",     rate,           C_ORANGE, "FFEDD5"),
    ]
    for i, (label, val, txt, bg) in enumerate(cards):
        col = get_column_letter(2 + i)
        ws2.column_dimensions[col].width = 20

        vc = ws2[f"{col}6"]
        vc.value     = val
        vc.fill      = _fill(bg)
        vc.font      = Font(bold=True, size=30, color=txt, name="Segoe UI")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[6].height = 56

        lc = ws2[f"{col}7"]
        lc.value     = label
        lc.fill      = _fill(txt)
        lc.font      = Font(bold=True, size=11, color=C_WHITE, name="Segoe UI")
        lc.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[7].height = 26

    # ── SHEET 3: By Area ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Area")
    ws3.sheet_view.showGridLines = False

    area_stats: dict[str, dict] = {}
    for obs in observations:
        area   = (obs.get("area") or "Unknown").strip()
        status = obs.get("status", "").strip().lower()
        if area not in area_stats:
            area_stats[area] = {"Total": 0, "Open": 0, "Partial": 0, "Closed": 0}
        area_stats[area]["Total"] += 1
        if status == "open":                            area_stats[area]["Open"]    += 1
        elif "partial" in status or "progress" in status: area_stats[area]["Partial"] += 1
        elif status == "closed":                        area_stats[area]["Closed"]  += 1

    ah = ["Area", "Total", "Open", "Partially Closed", "Closed", "Closure %"]
    for c, h in enumerate(ah, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.fill      = _fill(C_ORANGE)
        cell.font      = Font(bold=True, color=C_WHITE, name="Segoe UI", size=11)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = _border()
    ws3.row_dimensions[1].height = 28

    for r, (area, counts) in enumerate(area_stats.items(), 2):
        cl_pct = f"{counts['Closed'] / counts['Total'] * 100:.0f}%" if counts["Total"] else "N/A"
        rv = [area, counts["Total"], counts["Open"], counts["Partial"], counts["Closed"], cl_pct]
        bg = "F0F4FF" if r % 2 == 0 else "FFFFFF"
        for c, val in enumerate(rv, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.fill      = _fill(bg)
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center")
            cell.border    = _border()
            cell.font      = Font(name="Segoe UI", size=11)

    ws3.column_dimensions["A"].width = 38
    for i in range(2, 7):
        ws3.column_dimensions[get_column_letter(i)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
